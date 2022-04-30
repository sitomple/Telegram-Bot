import telebot
import openpyxl
from groups.ISP2 import isp_101_52_00
from groups.ISP2 import isp_202_52_00
from groups.ISP2 import isp_203_52_00
from groups.ISP2 import isp_204_52_00
from groups.ISP1 import isp_105_52_00
from groups.ISP1 import isp_104_52_00
from groups.ISP1 import isp_102_52_00
from groups.ISP1 import isp_103_52_00
from groups.ISP3 import isp_201_52_00
from groups.ISP3 import isp_302_52_00
from groups.ISP3 import isp_303_52_00
from groups.ISP3 import isp_304_52_00
from groups.ISP3 import isp_305_52_00
from groups.ISP4 import isp_401_301_52_00
from groups.ISP4 import isp_402_403_52_00
from groups.ZIO1 import zio_102_52_00
from groups.ZIO2 import zio_202_101_52_00
from groups.ZIO3 import zio_302_201_52_00
from groups.FK1 import fk_102_104_52_00
from groups.FK1 import fk_103_52_00
from groups.FK2 import fk_202_52_00
from groups.FK2 import fk_203_101_52_00
from groups.FK3 import fk_302_52_00
from groups.FK3 import fk_303_52_00
from groups.FK3 import fk_301_201_52_00
from groups.FK4 import fk_401_402_304_52_00
from groups.PSO1 import pso_103_104_52_00
from groups.PSO2 import pso_101_102_51_00
from groups.PSO2 import pso_203_52_00
from groups.PSO2 import pso_204_52_00
from groups.PSO3 import pso_201_202_51_00
from groups.PSO3 import pso_303_52_00
from groups.PSO3 import pso_304_52_00
from groups.EKN1 import ekn_102_52_00
from groups.FNK1 import fnk_102_52_00
from groups.EKN2 import ekn_101_202_52_00
from groups.FNK2 import fnk_202_52_00
from groups.FNK3 import fnk_302_201_52_00
import setings

from datetime import datetime, date, time

bot = telebot.TeleBot(setings.TOKEN_TELEGRAM)
wb = openpyxl.load_workbook(setings.PATH_TO_TABLE)
sheet = wb['СПО']

def day_lessons(message):
	wb = openpyxl.load_workbook('base.xlsx')
	sheet = wb['List']
	B = sheet.max_row
	cells = sheet['A1':f'B{B}']
	for id, gr in cells:
		if id.value == str(message.chat.id):
			if gr.value == 'ИСПк-101':
				less = isp_101_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-202':
				less = isp_202_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-203':
				less = isp_203_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-204':
				less = isp_204_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-104':
				less = isp_104_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-105':
				less = isp_105_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-102':
				less = isp_102_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-103':
				less = isp_103_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-201':
				less = isp_201_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-302':
				less = isp_302_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-303':
				less = isp_303_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-304':
				less = isp_304_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-305':
				less = isp_305_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-401':
				less = isp_401_301_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ИСПк-402':
				less = isp_402_403_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ЗИОк-102':
				less = zio_102_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ЗИОк-202':
				less = zio_202_101_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ЗИОк-302':
				less = zio_302_201_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-102':
				less = fk_102_104_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-103':
				less = fk_103_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-202':
				less = fk_202_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-203':
				less = fk_203_101_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-301':
				less = fk_301_201_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-302':
				less = fk_302_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-303':
				less = fk_303_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФКк-401':
				less = fk_401_402_304_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ПСОк-103':
				less = pso_103_104_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ПСОк-101':
				less = pso_101_102_51_00
				less.handle_text(message, bot)
			elif gr.value == 'ПСОк-203':
				less = pso_203_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ПСОк-204':
				less = pso_204_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ПСОк-201':
				less = pso_201_202_51_00
				less.handle_text(message, bot)
			elif gr.value == 'ПСОк-303':
				less = pso_303_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ПСОк-304':
				less = pso_304_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ЭКНк-102':
				less = ekn_102_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФНк-102':
				less = fnk_102_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ЭКНк-101':
				less = ekn_101_202_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФНк-202':
				less = fnk_202_52_00
				less.handle_text(message, bot)
			elif gr.value == 'ФНк-302':
				less = fnk_302_201_52_00
				less.handle_text(message, bot)


def base_group(bot, group, message):
		bot.send_message(message.chat.id, f'Активная группа: {message.text.strip()}-52-00')
		tumbler = 0
		i = 0
		wb = openpyxl.load_workbook('base.xlsx')
		sheet = wb['List']
		B = sheet.max_row
		cells = sheet['A1':f'B{B}']
		for id, gr in cells:
			i += 1
			if id.value == str(message.chat.id):
				tumbler = 1
				sheet[f'B{i}'] = f'{group}'
			wb.save('base.xlsx')
		if tumbler == 0:
			sheet[f'B{B + 1}'] = f'{group}'
			sheet[f'A{B + 1}'] = f'{message.chat.id}'
			wb.save('base.xlsx')

# Команда start
@bot.message_handler(commands=["start"])
def start(m, res=False):
		bot.send_message(m.chat.id, 'Бот был написан Zarnii и Sitomple\nДля групп ИСПк 2 курса \n'
									'Напишите Help')
		# Добавляем две кнопки
		markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
		item1 = telebot.types.KeyboardButton("Help")
		item2 = telebot.types.KeyboardButton("Сегодня")
		item3 = telebot.types.KeyboardButton("Завтра")
		item4 = telebot.types.KeyboardButton("Послезавтра")
		item5 = telebot.types.KeyboardButton("Неделя")
		markup.add(item1, item2, item3)
		markup.add(item4, item5)
		bot.send_message(m.chat.id, 'Выберите день, на который хотите получить расписание:',  reply_markup=markup)


@bot.message_handler(content_types=["text"])
def handle_text(message):
	if (message.text.strip().lower() == 'help') or (message.text.strip().lower() == 'расписание'):
		bot.send_message(message.chat.id, 'Напишите свою группу \nПример: \n*ИСПк-204 \n*ЗИОк-202 \n*ФКк-202 \n*ПСОк-204 \n*ЭКНк-102 \n*ФНк-202')

	if (message.text.strip() == 'ИСПк-101-51-00') or (message.text.strip() == 'ИСПк-101'):
		group = 'ИСПк-101'
		base_group(bot, group, message)

	elif (message.text.strip() == 'ИСПк-202-52-00') or (message.text.strip() == 'ИСПк-202'):
		group = 'ИСПк-202'
		base_group(bot, group, message)

	elif (message.text.strip() == 'ИСПк-203-52-00') or (message.text.strip() == 'ИСПк-203'):
		group = 'ИСПк-203'
		base_group(bot, group, message)

	elif (message.text.strip() == 'ИСПк-204-52-00') or (message.text.strip() == 'ИСПк-204'):
		group = 'ИСПк-204'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-105-52-00') or (message.text.strip() == 'ИСПк-105'):
		group = 'ИСПк-105'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-104-52-00') or (message.text.strip() == 'ИСПк-104'):
		group = 'ИСПк-104'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-102-52-00') or (message.text.strip() == 'ИСПк-102'):
		group = 'ИСПк-102'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-103-52-00') or (message.text.strip() == 'ИСПк-103'):
		group = 'ИСПк-103'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-201-52-00') or (message.text.strip() == 'ИСПк-201'):
		group = 'ИСПк-201'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-302-52-00') or (message.text.strip() == 'ИСПк-302'):
		group = 'ИСПк-302'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-303-52-00') or (message.text.strip() == 'ИСПк-303'):
		group = 'ИСПк-303'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-304-52-00') or (message.text.strip() == 'ИСПк-304'):
		group = 'ИСПк-304'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-305-52-00') or (message.text.strip() == 'ИСПк-305'):
		group = 'ИСПк-305'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-301-51-00') or (message.text.strip() == 'ИСПк-401-52-00') or (message.text.strip() == 'ИСПк-301') or (message.text.strip() == 'ИСПк-401') :
		group = 'ИСПк-401'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ИСПк-402-51-00') or (message.text.strip() == 'ИСПк-403-52-00') or (message.text.strip() == 'ИСПк-402') or (message.text.strip() == 'ИСПк-403'):
		group = 'ИСПк-402'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ЗИОк-102-52-00') or (message.text.strip() == 'ЗИОк-102'):
		group = 'ЗИОк-102'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ЗИОк-202-51-00') or (message.text.strip() == 'ЗИОк-101-51-00') or (message.text.strip() == 'ЗИОк-202') or (message.text.strip() == 'ЗИОк-101'):
		group = 'ЗИОк-202'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ЗИОк-302-51-00') or (message.text.strip() == 'ЗИОк-201-51-00') or (message.text.strip() == 'ЗИОк-302') or (message.text.strip() == 'ЗИОк-101'):
		group = 'ЗИОк-302'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-102-52-00') or (message.text.strip() == 'ФКк-104-51-00') or (message.text.strip() == 'ФКк-102') or (message.text.strip() == 'ФКк-104'):
		group = 'ФКк-102'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-103-52-00') or (message.text.strip() == 'ФКк-103'):
		group = 'ФКк-103'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-202-52-00') or (message.text.strip() == 'ФКк-202'):
		group = 'ФКк-202'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-203-52-00') or (message.text.strip() == 'ФКк-101-51-00') or (message.text.strip() == 'ФКк-203') or (message.text.strip() == 'ФКк-101'):
		group = 'ФКк-203'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-301-52-00') or (message.text.strip() == 'ФКк-201-51-00') or (message.text.strip() == 'ФКк-301') or (message.text.strip() == 'ФКк-201'):
		group = 'ФКк-301'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-302-52-00') or (message.text.strip() == 'ФКк-302'):
		group = 'ФКк-302'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-303-52-00') or (message.text.strip() == 'ФКк-303'):
		group = 'ФКк-303'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФКк-401-52-00') or (message.text.strip() == 'ФКк-402-52-00') or (message.text.strip() == 'ФКк-304-52-00') or (message.text.strip() == 'ФКк-401') or (message.text.strip() == 'ФКк-402') or (message.text.strip() == 'ФКк-304'):
		group = 'ФКк-401'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ПСОк-103-52-00') or (message.text.strip() == 'ПСОк-104-52-00') or (message.text.strip() == 'ПСОк-103') or (message.text.strip() == 'ПСОк-104'):
		group = 'ПСОк-103'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ПСОк-101-51-00') or (message.text.strip() == 'ПСОк-102-51-00') or (message.text.strip() == 'ПСОк-101') or (message.text.strip() == 'ПСОк-102'):
		group = 'ПСОк-101'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ПСОк-203-52-00') or (message.text.strip() == 'ПСОк-203'):
		group = 'ПСОк-203'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ПСОк-204-52-00') or (message.text.strip() == 'ПСОк-204'):
		group = 'ПСОк-204'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ПСОк-303-52-00') or (message.text.strip() == 'ПСОк-303'):
		group = 'ПСОк-303'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ПСОк-304-52-00') or (message.text.strip() == 'ПСОк-304'):
		group = 'ПСОк-304'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ПСОк-201-51-00') or (message.text.strip() == 'ПСОк-202-51-00') or (message.text.strip() == 'ПСОк-201') or (message.text.strip() == 'ПСОк-202'):
		group = 'ПСОк-201'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ЭКНк-102-52-00') or (message.text.strip() == 'ЭКНк-102'):
		group = 'ЭКНк-102'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФНк-102-52-00') or (message.text.strip() == 'ФНк-102'):
		group = 'ФНк-102'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ЭКНк-101-51-00') or (message.text.strip() == 'ЭКНк-202-52-00') or (message.text.strip() == 'ЭКНк-101') or (message.text.strip() == 'ЭКНк-202'):
		group = 'ЭКНк-101'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФНк-202-52-00') or (message.text.strip() == 'ФНк-202'):
		group = 'ФНк-202'
		base_group(bot, group, message)
	elif (message.text.strip() == 'ФНк-201-51-00') or (message.text.strip() == 'ФНк-302-52-00') or (message.text.strip() == 'ФНк-201') or (message.text.strip() == 'ФНк-302'):
		group = 'ФНк-302'
		base_group(bot, group, message)



	if message.text.strip() == 'Сегодня':
		day_lessons(message)

	elif message.text.strip() == 'Завтра':
		day_lessons(message)

	if message.text.strip() == 'Послезавтра':
		day_lessons(message)

	if message.text.strip() == 'Неделя':
		day_lessons(message)

bot.infinity_polling()